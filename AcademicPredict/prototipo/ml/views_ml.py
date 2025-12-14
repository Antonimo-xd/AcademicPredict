from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Avg
from django.core.management import call_command
from django.utils import timezone
import csv
import logging
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from prototipo.models import (
    PrediccionDesercionUniversitaria, 
    RegistroAcademicoUniversitario,
    TrazabilidadPrediccionDesercion,
    AlertaEstudiante,
    FichaSeguimientoEstudiante
)

logger = logging.getLogger(__name__)

@login_required
@user_passes_test(lambda u: hasattr(u, 'perfil'))
def dashboard_ml(request):
    """Dashboard principal de Machine Learning con Filtro Inteligente."""
    
    user_rol = request.user.perfil.rol

    predicciones = PrediccionDesercionUniversitaria.objects.all()
    
    # FILTRO POR ROL
    if user_rol == 'analista':
        # Solo predicciones de estudiantes asignados al analista
        predicciones = predicciones.filter(
            estudiante__alertas__analista_asignado=request.user
        ).distinct()
    
    elif user_rol == 'coordinador_carrera':
        # Solo predicciones de estudiantes de su carrera
        carrera = request.user.perfil.carrera_asignada
        if carrera:
            predicciones = predicciones.filter(estudiante__carrera=carrera)
        else:
            messages.error(request, "No tienes carrera asignada")
            return redirect('home')

    nivel_riesgo_filtro = request.GET.get('nivel_riesgo', '')
    busqueda = request.GET.get('busqueda', '')
    solo_anomalias = request.GET.get('anomalias', '')
    
    todas_predicciones = PrediccionDesercionUniversitaria.objects.all()
    
    criticos = todas_predicciones.filter(nivel_riesgo='Critico').count()
    altos = todas_predicciones.filter(nivel_riesgo='Alto').count()
    medios = todas_predicciones.filter(nivel_riesgo='Medio').count()
    bajos = todas_predicciones.filter(nivel_riesgo='Bajo').count()
    total_estudiantes = todas_predicciones.count()
    
    pct_criticos = (criticos / total_estudiantes * 100) if total_estudiantes > 0 else 0
    pct_altos = (altos / total_estudiantes * 100) if total_estudiantes > 0 else 0
    pct_medios = (medios / total_estudiantes * 100) if total_estudiantes > 0 else 0
    pct_bajos = (bajos / total_estudiantes * 100) if total_estudiantes > 0 else 0 
    
    ultima_prediccion = todas_predicciones.order_by('-fecha_prediccion').first()
    ultima_fecha = ultima_prediccion.fecha_prediccion if ultima_prediccion else None
    
    predicciones = PrediccionDesercionUniversitaria.objects.select_related(
        'estudiante',
        'estudiante__carrera',
        'registro_academico'
    )
    
    if nivel_riesgo_filtro:
        predicciones = predicciones.filter(nivel_riesgo=nivel_riesgo_filtro)
    
    if busqueda:
        estudiantes_ids = list(
            predicciones.filter(
                estudiante__codigo_estudiante__icontains=busqueda
            ).values_list('estudiante_id', flat=True).distinct()
        )
        predicciones = predicciones.filter(estudiante_id__in=estudiantes_ids)
    
    if solo_anomalias == '1':
        predicciones = predicciones.filter(es_anomalia=True)
    
    orden_riesgo = {
        'Critico': 1,
        'Alto': 2,
        'Medio': 3,
        'Bajo': 4
    }
    predicciones_lista = list(predicciones)
    predicciones_lista.sort(key=lambda x: (orden_riesgo.get(x.nivel_riesgo, 5), -x.probabilidad_desercion))
    
    paginator = Paginator(predicciones_lista, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'criticos': criticos,
        'altos': altos,
        'medios': medios,
        'bajos': bajos,
        'total_estudiantes': total_estudiantes,

        'pct_criticos': pct_criticos,
        'pct_altos': pct_altos,
        'pct_medios': pct_medios,
        'pct_bajos': pct_bajos,
        
        'ultima_prediccion': ultima_fecha,
        'page_obj': page_obj,
        
        'nivel_riesgo_actual': nivel_riesgo_filtro,
        'busqueda_actual': busqueda,
        'solo_anomalias_actual': solo_anomalias,

        'user_rol': user_rol,  
        'puede_ejecutar_ml': user_rol == 'admin',  
        'puede_asignar': user_rol in ['admin', 'coordinador'], 
    
    }
    
    return render(request, 'ml/dashboard_ml.html', context)

@login_required
@user_passes_test(lambda u: hasattr(u, 'perfil'))
def estudiante_detalle_ml(request, estudiante_id):

    prediccion = get_object_or_404(
        PrediccionDesercionUniversitaria,
        estudiante_id=estudiante_id
    )
    estudiante = prediccion.estudiante
    
    user_rol = request.user.perfil.rol

    if user_rol == 'coordinador_carrera':
        # Solo puede ver estudiantes de su carrera
        if estudiante.carrera != request.user.perfil.carrera_asignada:
            messages.error(request, "⛔ No tienes acceso a este estudiante de otra carrera")
            return redirect('dashboard_ml')
    
    elif user_rol == 'analista':
        # Solo puede ver estudiantes asignados a él
        from prototipo.models import AlertaEstudiante
        alerta = AlertaEstudiante.objects.filter(
            estudiante=estudiante,
            analista_asignado=request.user
        ).first()
        
        if not alerta:
            messages.error(request, "⛔ Este estudiante no está asignado a ti")
            return redirect('dashboard_ml')

    promedio_cohorte = PrediccionDesercionUniversitaria.objects.filter(
        estudiante__carrera=estudiante.carrera,
        estudiante__anio_ingreso_universidad=estudiante.anio_ingreso_universidad
    ).aggregate(promedio=Avg('probabilidad_desercion'))['promedio']
    
    if promedio_cohorte is None:
        promedio_cohorte = PrediccionDesercionUniversitaria.objects.aggregate(
            promedio=Avg('probabilidad_desercion')
        )['promedio']
    
    promedio_cohorte_pct = (promedio_cohorte * 100) if promedio_cohorte else 0
    
    estudiantes_cohorte = PrediccionDesercionUniversitaria.objects.filter(
        estudiante__carrera=estudiante.carrera,
        estudiante__anio_ingreso_universidad=estudiante.anio_ingreso_universidad
    )
    
    total_cohorte = estudiantes_cohorte.count()
    
    peores_que_este = estudiantes_cohorte.filter(
        probabilidad_desercion__gt=prediccion.probabilidad_desercion
    ).count()
    
    ranking = peores_que_este + 1
    
    total_universidad = PrediccionDesercionUniversitaria.objects.count()
    
    estudiantes_con_mayor_riesgo = PrediccionDesercionUniversitaria.objects.filter(
        probabilidad_desercion__gt=prediccion.probabilidad_desercion
    ).count()
    
    top_percentil = ((estudiantes_con_mayor_riesgo + 1) / total_universidad) * 100 if total_universidad > 0 else 0

    ranking_absoluto = estudiantes_con_mayor_riesgo + 1
    
    recomendaciones = obtener_recomendaciones_accion(prediccion.nivel_riesgo)
    
    factores_riesgo = prediccion.factores_riesgo if prediccion.factores_riesgo else []

    registro_actual = prediccion.registro_academico

    registro_anterior = RegistroAcademicoUniversitario.objects.filter(
        estudiante=estudiante,
        anio_academico__lt=registro_actual.anio_academico
    ).order_by('-anio_academico').first()

    if registro_anterior:
        if registro_actual.tareas_entregadas_total > (registro_anterior.tareas_entregadas_total * 1.1):
            tendencia_tareas = 'mejorando'
        elif registro_actual.tareas_entregadas_total < (registro_anterior.tareas_entregadas_total * 0.9):
            tendencia_tareas = 'empeorando'
        else:
            tendencia_tareas = 'estable'
        
        if registro_actual.dias_wifi_total > (registro_anterior.dias_wifi_total * 1.1):
            tendencia_asistencia = 'mejorando'
        elif registro_actual.dias_wifi_total < (registro_anterior.dias_wifi_total * 0.9):
            tendencia_asistencia = 'empeorando'
        else:
            tendencia_asistencia = 'estable'
        
        if registro_actual.visitas_lms_total > (registro_anterior.visitas_lms_total * 1.1):
            tendencia_lms = 'mejorando'
        elif registro_actual.visitas_lms_total < (registro_anterior.visitas_lms_total * 0.9):
            tendencia_lms = 'empeorando'
        else:
            tendencia_lms = 'estable'
    else:
        tendencia_tareas = 'sin_datos'
        tendencia_asistencia = 'sin_datos'
        tendencia_lms = 'sin_datos'

    tendencias = {
        'tareas': tendencia_tareas,
        'asistencia': tendencia_asistencia,
        'lms': tendencia_lms,
    }

    context = {
        'prediccion': prediccion,
        'estudiante': estudiante,
        'promedio_cohorte': promedio_cohorte_pct,
        'ranking': ranking,
        'total_cohorte': total_cohorte,
        'top_percentil': top_percentil,
        'recomendaciones': recomendaciones,
        'factores_riesgo': factores_riesgo,
        'tendencias': tendencias,
        'ranking_absoluto': ranking_absoluto,
        'total_universidad': total_universidad,
    }
    
    return render(request, 'ml/estudiante_detalle_ml.html', context)

def obtener_recomendaciones_accion(nivel_riesgo):
    """Retorna acciones recomendadas según nivel de riesgo"""
    
    recomendaciones = {
        'Critico': [
            '🚨 URGENTE: Contactar al estudiante en las próximas 24 horas',
            '📞 Agendar reunión presencial con tutor académico',
            '👥 Notificar inmediatamente al coordinador de carrera',
            '💼 Evaluar necesidad de apoyo financiero o becas',
            '🎯 Asignar tutor par para acompañamiento constante',
            '📊 Revisar historial académico completo y detectar patrones',
            '🏥 Considerar derivación a apoyo psicopedagógico',
            '📝 Elaborar plan de acción personalizado con metas semanales',
            '⏰ Establecer seguimiento diario durante 2 semanas',
            '📧 Enviar email formal de seguimiento con recursos disponibles'
        ],
        'Alto': [
            '📧 Contactar al estudiante en las próximas 48 horas',
            '📚 Derivar a tutorías académicas en materias críticas',
            '🔍 Realizar seguimiento semanal de asistencia y entregas',
            '📝 Revisar asignaturas con mayor dificultad',
            '🤝 Ofrecer apoyo psicopedagógico si es necesario',
            '💡 Proporcionar estrategias de estudio y organización',
            '👥 Invitar a grupos de estudio o talleres de reforzamiento',
            '📊 Monitorear actividad en plataforma LMS semanalmente',
            '🎓 Informar sobre recursos académicos disponibles',
            '📅 Agendar revisión de progreso en 2 semanas'
        ],
        'Medio': [
            '📅 Establecer seguimiento quincenal',
            '📨 Enviar email con recursos académicos disponibles',
            '📖 Promover uso de biblioteca y recursos digitales',
            '👀 Monitorear actividad en plataforma LMS',
            '🎯 Sugerir técnicas de estudio efectivas',
            '📊 Revisar rendimiento en próximas evaluaciones',
            '🤝 Ofrecer participación en talleres de desarrollo académico',
            '💬 Mantener canal de comunicación abierto',
            '📈 Reforzar positivamente los logros actuales',
            '🔔 Recordar fechas importantes de evaluaciones'
        ],
        'Bajo': [
            '✅ Mantener monitoreo mensual preventivo',
            '🎉 Reconocer y reforzar buen desempeño académico',
            '📢 Invitar a participar en actividades extracurriculares',
            '👥 Considerar como tutor par para otros estudiantes',
            '💡 Compartir mejores prácticas de estudio',
            '🏆 Postular a programas de excelencia académica',
            '📚 Sugerir material avanzado o proyectos adicionales',
            '🌟 Incluir en programas de liderazgo estudiantil',
            '📊 Mantener comunicación semestral',
            '✨ Reconocer públicamente logros académicos'
        ]
    }
    
    return recomendaciones.get(nivel_riesgo, [
        '📊 Realizar evaluación personalizada del caso',
        '👥 Consultar con coordinador académico',
        '📝 Documentar situación y definir plan de acción'
    ])

@login_required
@user_passes_test(lambda u: u.perfil.rol == 'admin', login_url='/sin-permiso/')
def ejecutar_deteccion_api(request):
    """Endpoint API para ejecutar detección ML desde el dashboard."""
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        output = StringIO()
        
        logger.info("🚀 Iniciando detección ML desde dashboard...")
        
        # Llamar al comando Django
        call_command('ejecutar_deteccion_ml', stdout=output)
        
        # Generar alertas basadas en nuevas predicciones
        alertas_generadas = generar_alertas_desde_predicciones()

        # Obtener estadísticas actualizadas
        total = PrediccionDesercionUniversitaria.objects.count()
        criticos = PrediccionDesercionUniversitaria.objects.filter(nivel_riesgo='Critico').count()
        altos = PrediccionDesercionUniversitaria.objects.filter(nivel_riesgo='Alto').count()
        medios = PrediccionDesercionUniversitaria.objects.filter(nivel_riesgo='Medio').count()
        
        logger.info(f"✅ Detección completada: {total} estudiantes procesados")
        logger.info(f"🔔 Alertas generadas: {alertas_generadas}")
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Detección ML completada exitosamente',
            'estudiantes_procesados': total,
            'criticos': criticos,
            'altos': altos,
            'medios': medios,
            'alertas_generadas': alertas_generadas,
            'fecha': timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        })
        
    except Exception as e:
        logger.error(f"❌ Error ejecutando detección ML: {e}")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'mensaje': 'Error al ejecutar detección ML'
        }, status=500)

def generar_alertas_desde_predicciones():
    alertas_creadas = 0
    
    # Obtener predicciones que requieren alertas (Alto y Crítico)
    predicciones_criticas = PrediccionDesercionUniversitaria.objects.filter(
        nivel_riesgo__in=['Alto', 'Critico']
    ).select_related('estudiante')
    
    for pred_uni in predicciones_criticas:
        try:
            estudiante = pred_uni.estudiante
            
            # Convertir nivel_riesgo a clasificacion
            clasificacion_mapa = {
                'Critico': 'critico',
                'Alto': 'alto',
                'Medio': 'medio',
                'Bajo': 'bajo'
            }
            clasificacion = clasificacion_mapa.get(pred_uni.nivel_riesgo, 'medio')
            
            # 1. Crear TrazabilidadPrediccionDesercion (registro histórico)
            trazabilidad = TrazabilidadPrediccionDesercion.objects.create(
                estudiante=estudiante,
                probabilidad_desercion=pred_uni.probabilidad_desercion,
                indice_riesgo=pred_uni.probabilidad_desercion * 100,
                clasificacion_riesgo=clasificacion,
                factores_riesgo=[],  # Puedes agregar factores si los tienes
                modelo_usado='XGBoost',
                version_modelo='1.0',
                activa=True
            )
            
            # 2. Crear AlertaEstudiante
            prioridad = 'critica' if clasificacion == 'critico' else 'alta'
            
            # Verificar si ya existe alerta pendiente para evitar duplicados
            alerta_existente = AlertaEstudiante.objects.filter(
                estudiante=estudiante,
                estado__in=['pendiente', 'en_revision'],
                visible=True
            ).exists()
            
            if not alerta_existente:
                AlertaEstudiante.objects.create(
                    estudiante=estudiante,
                    prediccion=trazabilidad,
                    tipo_alerta='riesgo_alto',
                    prioridad=prioridad,
                    titulo=f"Estudiante en Riesgo {pred_uni.nivel_riesgo}",
                    mensaje=f"El modelo ML detectó probabilidad de deserción de {pred_uni.probabilidad_desercion:.1%}",
                    indice_riesgo_momento=pred_uni.probabilidad_desercion * 100,
                    estado='pendiente'
                )
                
                # 3. Actualizar/crear FichaSeguimientoEstudiante
                ficha, created = FichaSeguimientoEstudiante.objects.get_or_create(
                    estudiante=estudiante
                )
                ficha.en_seguimiento = True
                ficha.ultimo_indice_riesgo = pred_uni.probabilidad_desercion * 100
                ficha.ultima_clasificacion = clasificacion
                ficha.ultima_fecha_prediccion = timezone.now()
                
                # Actualizar contadores
                ficha.alertas_activas = estudiante.alertas.filter(
                    estado__in=['pendiente', 'en_revision'],
                    visible=True
                ).count()
                ficha.save()
                
                alertas_creadas += 1
                logger.info(f"✅ Alerta generada: {estudiante.codigo_estudiante} - {prioridad.upper()}")
        
        except Exception as e:
            logger.error(f"❌ Error generando alerta para {pred_uni.estudiante.codigo_estudiante}: {e}")
            continue
    
    return alertas_creadas

@login_required
@user_passes_test(lambda u: hasattr(u, 'perfil'))
def exportar_csv(request):
    """Exporta predicciones a archivo CSV."""
    
    user_rol = request.user.perfil.rol
    nivel_riesgo = request.GET.get('nivel_riesgo', '')
    solo_anomalias = request.GET.get('anomalias', '')
    
    # ✅ CORRECCIÓN: Primero crear el queryset base
    predicciones = PrediccionDesercionUniversitaria.objects.select_related(
        'estudiante',
        'registro_academico'
    ).order_by('-probabilidad_desercion')
    
    if user_rol == 'analista':
        predicciones = predicciones.filter(
            estudiante__alertas__analista_asignado=request.user
        ).distinct()
    elif user_rol == 'coordinador_carrera':
        carrera = request.user.perfil.carrera_asignada
        if carrera:
            predicciones = predicciones.filter(estudiante__carrera=carrera)
    
    # Aplicar filtros de búsqueda
    if nivel_riesgo:
        predicciones = predicciones.filter(nivel_riesgo=nivel_riesgo)
    
    if solo_anomalias == '1':
        predicciones = predicciones.filter(es_anomalia=True)
    
    # El resto queda igual
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="estudiantes_riesgo.csv"'
    
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    writer.writerow([
        'ID Estudiante',
        'Código',
        'Probabilidad Deserción (%)',
        'Nivel Riesgo',
        'Es Anomalía',
        'Score Anomalía',
        'Rendimiento Predicho (%)',
        'Cantidad Factores Riesgo',
        'Factores Principales',
        'Fecha Predicción',
        'Año Académico'
    ])
    
    for pred in predicciones:
        factores_top = pred.factores_riesgo[:3] if pred.factores_riesgo else []
        factores_str = '; '.join([f.get('factor', '') for f in factores_top])
        
        writer.writerow([
            pred.estudiante.id,
            pred.estudiante.codigo_estudiante or '',
            f"{pred.probabilidad_desercion * 100:.2f}",
            pred.nivel_riesgo,
            'Sí' if pred.es_anomalia else 'No',
            f"{pred.score_anomalia:.4f}" if pred.score_anomalia else '',
            f"{pred.rendimiento_predicho_futuro:.2f}" if pred.rendimiento_predicho_futuro else '',
            len(pred.factores_riesgo) if pred.factores_riesgo else 0,
            factores_str,
            pred.fecha_prediccion.strftime('%d/%m/%Y %H:%M'),
            pred.registro_academico.anio_academico if pred.registro_academico else ''
        ])
    
    return response

@login_required
@user_passes_test(lambda u: hasattr(u, 'perfil'))
def exportar_excel(request):
    """Exporta predicciones a archivo Excel con formato."""
    
    user_rol = request.user.perfil.rol
    nivel_riesgo = request.GET.get('nivel_riesgo', '')
    
    predicciones = PrediccionDesercionUniversitaria.objects.select_related(
        'estudiante',
        'registro_academico'
    ).order_by('-probabilidad_desercion')
    
    if user_rol == 'analista':
        predicciones = predicciones.filter(
            estudiante__alertas__analista_asignado=request.user
        ).distinct()
    elif user_rol == 'coordinador_carrera':
        carrera = request.user.perfil.carrera_asignada
        if carrera:
            predicciones = predicciones.filter(estudiante__carrera=carrera)

    if nivel_riesgo:
        predicciones = predicciones.filter(nivel_riesgo=nivel_riesgo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Predicciones ML"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = [
        'ID', 'Código', 'Probabilidad (%)', 
        'Nivel Riesgo', 'Anomalía', 'Factores', 'Fecha'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row, pred in enumerate(predicciones, 2):
        ws.cell(row=row, column=1, value=pred.estudiante.id)
        ws.cell(row=row, column=2, value=pred.estudiante.codigo_estudiante or '')
        ws.cell(row=row, column=3, value=round(pred.probabilidad_desercion * 100, 2))
        
        nivel_cell = ws.cell(row=row, column=4, value=pred.nivel_riesgo)
        
        if pred.nivel_riesgo == 'Critico':
            nivel_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            nivel_cell.font = Font(color="FFFFFF", bold=True)
        elif pred.nivel_riesgo == 'Alto':
            nivel_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        elif pred.nivel_riesgo == 'Medio':
            nivel_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        else:
            nivel_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        ws.cell(row=row, column=5, value='Sí' if pred.es_anomalia else 'No')
        ws.cell(row=row, column=6, value=len(pred.factores_riesgo) if pred.factores_riesgo else 0)
        ws.cell(row=row, column=7, value=pred.fecha_prediccion.strftime('%d/%m/%Y'))
    
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="predicciones_ml.xlsx"'
    
    wb.save(response)
    
    return response

def calcular_tendencias(registros_historicos):
    """Calcula tendencias en el historial académico."""
    
    if len(registros_historicos) < 2:
        return {
            'tareas': 'sin_datos',
            'asistencia': 'sin_datos',
            'lms': 'sin_datos'
        }
    
    primero = registros_historicos.first()
    ultimo = registros_historicos.last()
    
    def calcular_tendencia(valor_inicial, valor_final):
        if valor_inicial == 0:
            return 'sin_datos'
        
        cambio_pct = ((valor_final - valor_inicial) / valor_inicial) * 100
        
        if cambio_pct > 10:
            return 'mejorando'  # ↑
        elif cambio_pct < -10:
            return 'empeorando'  # ↓
        else:
            return 'estable'  # →
    
    return {
        'tareas': calcular_tendencia(
            primero.tareas_entregadas_total or 0,
            ultimo.tareas_entregadas_total or 0
        ),
        'asistencia': calcular_tendencia(
            primero.dias_wifi_total or 0,
            ultimo.dias_wifi_total or 0
        ),
        'lms': calcular_tendencia(
            primero.visitas_lms_total or 0,
            ultimo.visitas_lms_total or 0
        )
    }

@login_required
def exportar_detalle_excel(request, estudiante_id):
    """ Exporta el detalle completo de un estudiante a Excel."""
    
    prediccion = get_object_or_404(
        PrediccionDesercionUniversitaria.objects.select_related(
            'estudiante',
            'registro_academico'
        ),estudiante_id=estudiante_id)
    
    estudiante = prediccion.estudiante
    
    user_rol = request.user.perfil.rol
    
    # Verificar acceso según rol
    if user_rol == 'coordinador_carrera':
        # Solo puede exportar estudiantes de su carrera
        if estudiante.carrera != request.user.perfil.carrera_asignada:
            messages.error(request, "⛔ No tienes acceso a este estudiante de otra carrera")
            return redirect('dashboard_ml')
    
    elif user_rol == 'analista':
        # Solo puede exportar estudiantes asignados a él
        from prototipo.models import AlertaEstudiante
        alerta = AlertaEstudiante.objects.filter(
            estudiante=estudiante,
            analista_asignado=request.user
        ).first()
        
        if not alerta:
            messages.error(request, "⛔ Este estudiante no está asignado a ti")
            return redirect('dashboard_ml')

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Estudiante"
    
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1)
    cell.value = f"REPORTE DE RIESGO - ESTUDIANTE {estudiante.codigo_estudiante}"
    cell.font = title_font
    cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 2
    ws.cell(row=row, column=1, value="Código:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=estudiante.codigo_estudiante)
    
    row += 1
    ws.cell(row=row, column=1, value="Carrera:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=estudiante.carrera.nombre)
    
    row += 1
    ws.cell(row=row, column=1, value="Año Académico:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=prediccion.registro_academico.anio_academico)
    
    row += 1
    ws.cell(row=row, column=1, value="Fecha Predicción:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=prediccion.fecha_prediccion.strftime('%d/%m/%Y %H:%M'))
    
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1)
    cell.value = "MÉTRICAS DE RIESGO"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    headers = ["Métrica", "Valor", "Interpretación", "Nivel"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.border = border
    
    row += 1
    prob_pct = prediccion.probabilidad_desercion * 100
    ws.cell(row=row, column=1, value="Probabilidad Deserción").border = border
    ws.cell(row=row, column=2, value=f"{prob_pct:.1f}%").border = border
    if prob_pct > 80:
        ws.cell(row=row, column=3, value="MUY ALTO RIESGO").border = border
        ws.cell(row=row, column=4, value="CRÍTICO").fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    elif prob_pct > 60:
        ws.cell(row=row, column=3, value="ALTO RIESGO").border = border
        ws.cell(row=row, column=4, value="ALTO").fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    else:
        ws.cell(row=row, column=3, value="Riesgo moderado").border = border
        ws.cell(row=row, column=4, value="MEDIO").fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
    ws.cell(row=row, column=4).border = border
    
    row += 1
    ws.cell(row=row, column=1, value="Nivel de Riesgo").border = border
    ws.cell(row=row, column=2, value=prediccion.nivel_riesgo).border = border
    ws.cell(row=row, column=3, value="Clasificación ML").border = border
    ws.cell(row=row, column=4, value=prediccion.nivel_riesgo).border = border
    
    row += 1
    ws.cell(row=row, column=1, value="Es Anomalía").border = border
    ws.cell(row=row, column=2, value="Sí" if prediccion.es_anomalia else "No").border = border
    ws.cell(row=row, column=3, value="Patrón atípico" if prediccion.es_anomalia else "Patrón normal").border = border
    ws.cell(row=row, column=4, value="⚠️" if prediccion.es_anomalia else "✓").border = border
    
    if prediccion.factores_riesgo:
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = "FACTORES DE RIESGO IDENTIFICADOS"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        headers = ["Factor", "Valor Actual", "Valor Esperado", "Impacto", "Peso (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.border = border
        
        for factor in prediccion.factores_riesgo:
            row += 1
            ws.cell(row=row, column=1, value=factor['factor']).border = border
            ws.cell(row=row, column=2, value=str(factor['valor_actual'])).border = border
            ws.cell(row=row, column=3, value=factor['valor_esperado']).border = border
            ws.cell(row=row, column=4, value=factor['impacto']).border = border
            ws.cell(row=row, column=5, value=f"{factor['peso']:.1f}%").border = border
            
            if factor['impacto'] == 'ALTO':
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = f"💡 Recomendación: {factor['recomendacion']}"
            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="detalle_estudiante_{estudiante.codigo_estudiante}.xlsx"'
    
    wb.save(response)
    
    return response


