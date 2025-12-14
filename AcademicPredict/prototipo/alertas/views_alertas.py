from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
import openpyxl
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime

from prototipo.models import (
    AlertaEstudiante,
    FichaSeguimientoEstudiante,
    IntervencionEstudiante,
    TrazabilidadPrediccionDesercion,
    EstudianteUniversitario
)


# ================================================================
# DASHBOARD DE ALERTAS
# ================================================================

@login_required
def dashboard(request):
    """
    Dashboard principal de alertas
    Muestra KPIs y listado filtrable de alertas
    """

    user_rol = request.user.perfil.rol
    
    if user_rol == 'analista':
        alertas = alertas.filter(analista_asignado=request.user)
    elif user_rol == 'coordinador_carrera':
        carrera = request.user.perfil.carrera_asignada
        if carrera:
            alertas = alertas.filter(estudiante__carrera=carrera)

    # Obtener filtros
    estado = request.GET.get('estado', '')
    prioridad = request.GET.get('prioridad', '')
    busqueda = request.GET.get('busqueda', '')
    
    # Query base
    alertas = AlertaEstudiante.objects.select_related(
        'estudiante', 
        'prediccion'
    ).order_by('-fecha_creacion')
    
    # Aplicar filtros
    if estado:
        alertas = alertas.filter(estado=estado)
    
    if prioridad:
        alertas = alertas.filter(prioridad=prioridad)
    
    if busqueda:
        alertas = alertas.filter(
            Q(estudiante__codigo_estudiante__icontains=busqueda)
        )
    
    # Calcular KPIs
    total_alertas = AlertaEstudiante.objects.count()
    alertas_criticas = AlertaEstudiante.objects.filter(prioridad='critica').count()
    alertas_altas = AlertaEstudiante.objects.filter(prioridad='alta').count()
    alertas_pendientes = AlertaEstudiante.objects.filter(estado='pendiente').count()
    estudiantes_seguimiento = FichaSeguimientoEstudiante.objects.filter(
        en_seguimiento=True
    ).count()
    
    # Calcular porcentajes
    pct_criticas = (alertas_criticas / total_alertas * 100) if total_alertas > 0 else 0
    pct_altas = (alertas_altas / total_alertas * 100) if total_alertas > 0 else 0
    
    # Paginación
    paginator = Paginator(alertas, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'alertas': page_obj,
        'alertas_criticas': alertas_criticas,
        'alertas_altas': alertas_altas,
        'alertas_pendientes': alertas_pendientes,
        'estudiantes_seguimiento': estudiantes_seguimiento,
        'pct_criticas': pct_criticas,
        'pct_altas': pct_altas,
        'page_obj': page_obj,
    }
    
    return render(request, 'alertas/dashboard_alertas.html', context)


@login_required
def detalle(request, pk):
    """Muestra el detalle completo de una alerta"""
    alerta = get_object_or_404(
        AlertaEstudiante.objects.select_related('estudiante', 'prediccion'),
        pk=pk
    )
    
    return render(request, 'alertas/detalle_alerta.html', {'alerta': alerta})


@login_required
@user_passes_test(lambda u: u.perfil.rol in ['admin', 'coordinador'])
def cambiar_estado(request, pk):
    """Cambia el estado de una alerta"""
    if request.method == 'POST':
        alerta = get_object_or_404(AlertaEstudiante, pk=pk)
        nuevo_estado = request.POST.get('estado')
        
        if nuevo_estado in ['pendiente', 'en_revision', 'resuelta', 'descartada']:
            alerta.estado = nuevo_estado
            alerta.save()
            messages.success(request, f'Estado actualizado a: {alerta.get_estado_display()}')
        else:
            messages.error(request, 'Estado inválido')
        
        return redirect('alertas:detalle', pk=pk)
    
    return redirect('alertas:dashboard')


@login_required
def seguimiento(request, codigo):
    """Ficha de seguimiento individual del estudiante"""
    estudiante = get_object_or_404(EstudianteUniversitario, codigo_estudiante=codigo)
    
    ficha, created = FichaSeguimientoEstudiante.objects.get_or_create(
        estudiante=estudiante
    )
    
    intervenciones = IntervencionEstudiante.objects.filter(
        estudiante=estudiante
    ).select_related('responsable').order_by('-fecha_intervencion')
    
    context = {
        'ficha': ficha,
        'intervenciones': intervenciones,
    }
    
    return render(request, 'alertas/ficha_seguimiento.html', context)


@login_required
@user_passes_test(lambda u: hasattr(u, 'perfil'))
def nueva_intervencion(request, codigo):
    """Registra una nueva intervención de seguimiento"""
    if request.method == 'POST':
        estudiante = get_object_or_404(EstudianteUniversitario, codigo_estudiante=codigo)
        
        ficha, created = FichaSeguimientoEstudiante.objects.get_or_create(
            estudiante=estudiante
        )
        
        # Crear intervención (sin campo ficha_seguimiento)
        IntervencionEstudiante.objects.create(
            estudiante=estudiante,
            tipo_intervencion=request.POST.get('tipo_intervencion'),
            titulo=f"{request.POST.get('tipo_intervencion')} - {estudiante.codigo_estudiante}",
            descripcion=request.POST.get('descripcion'),
            resultado=request.POST.get('resultado'),
            observaciones=request.POST.get('observaciones', ''),
            responsable=request.user
        )
        
        # Actualizar fecha de contacto en ficha
        ficha.ultima_fecha_contacto = timezone.now().date()
        ficha.save()
        
        messages.success(request, 'Intervención registrada exitosamente')
        return redirect('alertas:seguimiento', codigo=codigo)
    
    return redirect('alertas:seguimiento', codigo=codigo)


@login_required
@user_passes_test(lambda u: u.perfil.rol in ['admin', 'coordinador'])
def cambiar_estado_seguimiento(request, codigo):
    """Activa o desactiva el seguimiento de un estudiante"""
    if request.method == 'POST':
        estudiante = get_object_or_404(EstudianteUniversitario, codigo_estudiante=codigo)
        ficha = get_object_or_404(FichaSeguimientoEstudiante, estudiante=estudiante)
        
        en_seguimiento = request.POST.get('en_seguimiento') == '1'
        ficha.en_seguimiento = en_seguimiento
        ficha.ultima_actualizacion = timezone.now()
        ficha.save()
        
        estado_texto = 'activado' if en_seguimiento else 'desactivado'
        messages.success(request, f'Seguimiento {estado_texto} exitosamente')
        
        return redirect('alertas:seguimiento', codigo=codigo)
    
    return redirect('alertas:seguimiento', codigo=codigo)


@login_required
def listado_seguimiento(request):
    """Listado de todos los estudiantes en seguimiento"""
    
    busqueda = request.GET.get('busqueda', '')
    riesgo = request.GET.get('riesgo', '')
    intervenciones = request.GET.get('intervenciones', '')
    
    fichas = FichaSeguimientoEstudiante.objects.select_related(
        'estudiante'
    ).filter(en_seguimiento=True).order_by('-ultimo_indice_riesgo')
    
    if busqueda:
        fichas = fichas.filter(estudiante__codigo_estudiante__icontains=busqueda)
    
    if riesgo:
        if riesgo == 'critico':
            fichas = fichas.filter(ultimo_indice_riesgo__gte=70)
        elif riesgo == 'alto':
            fichas = fichas.filter(ultimo_indice_riesgo__gte=50, ultimo_indice_riesgo__lt=70)
        elif riesgo == 'medio':
            fichas = fichas.filter(ultimo_indice_riesgo__gte=30, ultimo_indice_riesgo__lt=50)
        elif riesgo == 'bajo':
            fichas = fichas.filter(ultimo_indice_riesgo__lt=30)
    
    fichas = fichas.annotate(
        num_intervenciones=Count('estudiante__intervenciones')  # Nombre diferente
    )

    if intervenciones == 'si':
        fichas = fichas.filter(num_intervenciones__gt=0)
    elif intervenciones == 'no':
        fichas = fichas.filter(num_intervenciones=0)
    
    # Estadísticas
    total_seguimiento = FichaSeguimientoEstudiante.objects.filter(en_seguimiento=True).count()
    riesgo_critico = FichaSeguimientoEstudiante.objects.filter(
        en_seguimiento=True, ultimo_indice_riesgo__gte=70
    ).count()
    riesgo_alto = FichaSeguimientoEstudiante.objects.filter(
        en_seguimiento=True, ultimo_indice_riesgo__gte=50, ultimo_indice_riesgo__lt=70
    ).count()
    con_intervenciones = FichaSeguimientoEstudiante.objects.filter(
        en_seguimiento=True
    ).annotate(total_int=Count('estudiante__intervenciones')).filter(total_int__gt=0).count()
    
    # Paginación
    paginator = Paginator(fichas, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'estudiantes': page_obj,
        'total_seguimiento': total_seguimiento,
        'riesgo_critico': riesgo_critico,
        'riesgo_alto': riesgo_alto,
        'con_intervenciones': con_intervenciones,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
    }
    
    return render(request, 'alertas/listado_seguimiento.html', context)


@login_required
def reportes(request):
    """Vista de generación de reportes"""
    
    total_alertas = AlertaEstudiante.objects.count()
    alertas_pendientes = AlertaEstudiante.objects.filter(estado='pendiente').count()
    alertas_resueltas = AlertaEstudiante.objects.filter(estado='resuelta').count()
    alertas_revision = AlertaEstudiante.objects.filter(estado='en_revision').count()
    alertas_descartadas = AlertaEstudiante.objects.filter(estado='descartada').count()
    
    estudiantes_seguimiento = FichaSeguimientoEstudiante.objects.filter(
        en_seguimiento=True
    ).count()
    
    alertas_criticas = AlertaEstudiante.objects.filter(prioridad='critica').count()
    alertas_altas = AlertaEstudiante.objects.filter(prioridad='alta').count()
    alertas_medias = AlertaEstudiante.objects.filter(prioridad='media').count()
    alertas_bajas = AlertaEstudiante.objects.filter(prioridad='baja').count()
    
    pct_criticas = (alertas_criticas / total_alertas * 100) if total_alertas > 0 else 0
    pct_altas = (alertas_altas / total_alertas * 100) if total_alertas > 0 else 0
    pct_medias = (alertas_medias / total_alertas * 100) if total_alertas > 0 else 0
    pct_bajas = (alertas_bajas / total_alertas * 100) if total_alertas > 0 else 0
    
    context = {
        'total_alertas': total_alertas,
        'alertas_pendientes': alertas_pendientes,
        'alertas_resueltas': alertas_resueltas,
        'alertas_revision': alertas_revision,
        'alertas_descartadas': alertas_descartadas,
        'estudiantes_seguimiento': estudiantes_seguimiento,
        'alertas_criticas': alertas_criticas,
        'alertas_altas': alertas_altas,
        'alertas_medias': alertas_medias,
        'alertas_bajas': alertas_bajas,
        'pct_criticas': pct_criticas,
        'pct_altas': pct_altas,
        'pct_medias': pct_medias,
        'pct_bajas': pct_bajas,
        'historial_reportes': [],
    }
    
    return render(request, 'alertas/reporte_seguimiento.html', context)


@login_required
def generar_reporte(request):
    """Genera un reporte según los parámetros seleccionados"""
    if request.method == 'POST':
        tipo_reporte = request.POST.get('tipo_reporte')
        formato = request.POST.get('formato')
        
        # TODO: Implementar generación de reportes
        messages.success(
            request, 
            f'Reporte {tipo_reporte} en formato {formato} generado exitosamente'
        )
        
        return redirect('alertas:reporte')
    
    return redirect('alertas:reporte')


@login_required
def api_alertas_urgentes(request):
    """
    API para obtener alertas urgentes
    Usado por el dropdown de alertas en el sidebar
    """
    
    alertas = AlertaEstudiante.objects.filter(
        estado__in=['pendiente', 'en_revision'],
        prioridad__in=['critica', 'alta']
    ).select_related('estudiante', 'prediccion').order_by(
        '-prioridad',
        '-fecha_creacion'
    )[:10]
    
    alertas_data = []
    for alerta in alertas:
        # Obtener índice de riesgo de la predicción (puede ser None)
        indice_riesgo = None
        if alerta.prediccion and alerta.prediccion.indice_riesgo is not None:
            indice_riesgo = float(alerta.prediccion.indice_riesgo)
        
        alertas_data.append({
            'id': alerta.id,
            'estudiante': alerta.estudiante.codigo_estudiante,
            'prioridad': alerta.get_prioridad_display(),
            'prioridad_codigo': alerta.prioridad,
            'titulo': alerta.titulo,
            'indice_riesgo': indice_riesgo, 
            'fecha_creacion': alerta.fecha_creacion.isoformat(),
        })
    
    return JsonResponse({
        'total': len(alertas_data),
        'alertas': alertas_data
    })

@login_required
def exportar_seguimiento(request):
    """Exporta listado de seguimiento a Excel"""
    
    # Obtener fichas en seguimiento
    fichas = FichaSeguimientoEstudiante.objects.select_related(
        'estudiante'
    ).filter(en_seguimiento=True).annotate(
        num_intervenciones=Count('estudiante__intervenciones')
    ).order_by('-ultimo_indice_riesgo')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seguimiento Estudiantes"
    
    # Headers
    headers = [
        'Código Estudiante',
        'Último Índice Riesgo (%)',
        'Clasificación',
        'Fecha Inicio Seguimiento',
        'Última Actualización',
        'Total Intervenciones',
        'Estado'
    ]
    ws.append(headers)
    
    # Datos
    for ficha in fichas:
        # Clasificación de riesgo
        if ficha.ultimo_indice_riesgo >= 70:
            clasificacion = 'Crítico'
        elif ficha.ultimo_indice_riesgo >= 50:
            clasificacion = 'Alto'
        elif ficha.ultimo_indice_riesgo >= 30:
            clasificacion = 'Medio'
        else:
            clasificacion = 'Bajo'
        
        ws.append([
            ficha.estudiante.codigo_estudiante,
            round(ficha.ultimo_indice_riesgo, 1) if ficha.ultimo_indice_riesgo else 0,
            clasificacion,
            ficha.fecha_inicio_seguimiento.strftime('%d/%m/%Y') if ficha.fecha_inicio_seguimiento else '',
            ficha.ultima_actualizacion.strftime('%d/%m/%Y %H:%M') if ficha.ultima_actualizacion else '',
            ficha.num_intervenciones,
            'Activo' if ficha.en_seguimiento else 'Inactivo'
        ])
    
    # Estilos
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename=seguimiento_estudiantes_{fecha}.xlsx'
    
    wb.save(response)
    return response




























